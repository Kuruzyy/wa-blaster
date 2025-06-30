#!/usr/bin/env python3
"""
HWUM WhatsApp Blaster - Refactored PySide6 Version with Batch Updates & Custom Placeholders

This module implements a dual-browser WhatsApp messaging tool with a GUI using PySide6.
Refactored for improved efficiency, reliability, readability, batch Excel updates,
and custom message placeholders.
"""

import os
import sys
import re
import shutil
import tempfile
import platform
import urllib.parse
import logging
import random
import time
import threading
import queue
import requests
import gdown
import mimetypes
from lxml import html
from pathlib import Path
from typing import Any, Dict, Optional, List, Tuple, Final

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

# --- GUI Framework ---
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
    QLabel, QLineEdit, QPushButton, QTextEdit, QCheckBox, QFileDialog, QMessageBox,
    QDialog, QProgressBar
)
from PySide6.QtCore import Qt, Signal, QObject, QThread
from PySide6.QtGui import QFont, QCloseEvent, QClipboard

# --- Selenium Imports ---
from selenium import webdriver
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    WebDriverException,
    ElementClickInterceptedException,
    StaleElementReferenceException,
    ElementNotInteractableException,
)

# --- Webdriver Managers ---
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.core.os_manager import ChromeType
from webdriver_manager.microsoft import EdgeChromiumDriverManager

# --- Constants ---
CONFIG: Dict[str, Any] = {
    "SHEETS": {
        "LIST": "LIST",
        "MSGS": "MSGS",
        "DOCS": "DOCS",
        "MEDIA": "MEDIA",
        "SETTINGS": "SETTINGS",
        "PLACEHOLDER": "PLACEHOLDER",
        "BROWSER": "BROWSER"
    },
    "COLUMNS": {
        "LIST": {
            "phone": "Phone Number",
            "msg_code": "Message Code",
            "doc_code": "Document Code",
            "media_code": "Media Code",
            "resolved_name": "_ResolvedName",
            "status": "Status",
        },
        "MSGS": {
            "msg_code": "Message Code",
            "message": "Message Encoded",
        },
        "DOCS": {
            "code": "Document Code",
            "files": ["BROCHURE_1", "BROCHURE_2", "BROCHURE_3", "BROCHURE_4"],
        },
        "MEDIA": {
            "code": "Media Code",
            "files": ["MEDIA_1", "MEDIA_2", "MEDIA_3", "MEDIA_4"],
        },
        "SETTINGS": {
            "use_browser": "USE_BROWSER",
            "custom_bsr_path": "BSR_PATH",
            "user_agent": "USER_AGENT",
            "wd_ver": "WD_VER",
            "xpath_text": "XPATH_TEXT",
            "xpath_send": "XPATH_SEND",
            "xpath_attach": "XPATH_ATTACH",
            "xpath_asend": "XPATH_ASEND",
            "xpath_docs": "XPATH_DOCS",
            "xpath_media": "XPATH_MEDIA",
            "invalid_message": "INVALID_MSG",
            "min_timer": "MIN_TIMER",
            "max_timer": "MAX_TIMER",
        },
        "PLACEHOLDER": {
            "keyword": "Keyword",
            "list_header": "LIST Name"
        },
        "BROWSER": {
            "name": "NAME",
            "paths": ["PATH_1", "PATH_2", "PATH_3", "PATH_4", "PATH_5", "PATH_6"]
        }
    },
    "STATUS_VALUES": {
        "PENDING": -1,
        "INVALID": 0,
        "SENT": 1,
        "RETRY": 2,
    },
    "DEFAULT_SETTINGS": {
        "USE_BROWSER": "CHROME",
        "BSR_PATH": None,
        "USER_AGENT": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36 Edg/110.0.1587.69",
        "WD_VER": None,
        "XPATH_TEXT": '/html/body/div[1]/div/div/div[3]/div/div[4]/div/footer/div[1]/div/span/div/div[2]/div[1]/div[2]/div[1]',
        "XPATH_SEND": '/html/body/div[1]/div/div/div[3]/div/div[4]/div/footer/div[1]/div/span/div/div[2]/div[2]/button',
        "XPATH_ATTACH": '/html/body/div[1]/div/div/div[3]/div/div[4]/div/footer/div[1]/div/span/div/div[1]/div/button',
        "XPATH_ASEND": '/html/body/div[1]/div/div/div[3]/div/div[2]/div[2]/span/div/div/div/div[2]/div/div[2]/div[2]/div/div',
        "XPATH_DOCS": '//*[@id="app"]/div/span[5]/div/ul/div/div/div[1]/li/div/span',
        "XPATH_MEDIA": '//*[@id="app"]/div/span[5]/div/ul/div/div/div[2]/li/div/span',
        "INVALID_MSG": "Phone number shared via url is invalid",
        "MIN_TIMER": "2.0",
        "MAX_TIMER": "5.0",
    }
}

# Define column names from CONFIG for easier access
_LIST_COLS = CONFIG["COLUMNS"]["LIST"]
_SETTINGS_COLS = CONFIG["COLUMNS"]["SETTINGS"]
_PLACEHOLDER_COLS = CONFIG["COLUMNS"]["PLACEHOLDER"]
_BROWSER_COLS = CONFIG["COLUMNS"]["BROWSER"]
_STATUS_VALS = CONFIG["STATUS_VALUES"]

# --- Logging Setup ---
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - [%(threadName)s] - %(message)s",
    handlers=[
        logging.FileHandler("whatsapp_blaster.log", encoding='utf-8'),
        logging.StreamHandler()
    ],
)

# --- Global Signals for GUI Updates from Threads ---
class Signals(QObject):
    log_system_signal = Signal(str)
    log_browser1_signal = Signal(str)
    log_browser2_signal = Signal(str)
    processing_started = Signal()
    processing_stopped = Signal()
    update_progress = Signal(int, int)

global_signals: Optional[Signals] = None

# --- Thread-Safe Logging Functions ---
def log_system(message: str) -> None:
    logging.info(message)
    if global_signals: global_signals.log_system_signal.emit(message)

def log_browser(instance_id: int, message: str) -> None:
    log_entry = f"[Instance {instance_id}] {message}"
    logging.info(log_entry)
    if global_signals:
        if instance_id == 1: global_signals.log_browser1_signal.emit(log_entry)
        else: global_signals.log_browser2_signal.emit(log_entry)

# --- Utility Functions ---
def get_persistent_temp_path(instance_id: str) -> Path:
    base = Path(tempfile.gettempdir())
    folder = f"whatsapp_blaster_data_{instance_id}"
    path = base / folder
    path.mkdir(parents=True, exist_ok=True)
    return path

def normalize_value(value: Any) -> str:
    if pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()

# --- Excel File Locking ---
excel_lock = threading.Lock()
def is_file_locked(filepath: str) -> bool:
    lock_path = Path(f"{filepath}.lock")
    if lock_path.exists():
        try:
            if (time.time() - lock_path.stat().st_mtime) > 60:
                log_system(f"Warning: Stale lock file found for {filepath}, removing.")
                lock_path.unlink()
                return False
            else:
                return True
        except OSError:
            return True
    try:
        open(lock_path, 'x').close()
        return False
    except FileExistsError: 
        return True
    except IOError as e:
        log_system(f"IOError checking lock for {filepath}: {e}")
        return True

def release_lock(filepath: str):
    lock_path = Path(f"{filepath}.lock")
    try:
        if lock_path.exists():
            lock_path.unlink()
    except OSError as e:
        log_system(f"Error removing lock file {lock_path}: {e}")

# --- Batch Excel Update Function ---
def perform_batch_update(excel_file: str, statuses_to_update: Dict[str, int]) -> bool:
    """Updates the Excel file status column in batch using openpyxl."""
    if not statuses_to_update:
        log_system("Batch update skipped: No statuses.")
        return True
    log_system(f"Starting batch update for {len(statuses_to_update)} statuses in '{excel_file}'...")
    max_attempts, retry_delay = 3, 2
    list_sheet_name = CONFIG["SHEETS"]["LIST"]
    phone_col_name, status_col_name = _LIST_COLS["phone"], _LIST_COLS["status"]
    success = False
    for attempt in range(max_attempts):
        if is_file_locked(excel_file):
            log_system(f"Excel file locked. Attempt {attempt + 1}/{max_attempts}. Retrying...")
            time.sleep(retry_delay)
            continue
        with excel_lock:
            try:
                wb = load_workbook(excel_file)
                if list_sheet_name not in wb.sheetnames:
                    log_system(f"Error: Sheet '{list_sheet_name}' not found. Batch update failed.")
                    release_lock(excel_file)
                    return False
                sheet = wb[list_sheet_name]
                headers = [cell.value for cell in sheet[1]]
                try:
                    phone_col_idx = headers.index(phone_col_name) + 1
                    status_col_idx = headers.index(status_col_name) + 1
                except ValueError:
                    log_system(f"Error: Columns '{phone_col_name}' or '{status_col_name}' not found. Batch update failed.")
                    wb.close()
                    release_lock(excel_file)
                    return False
                updated_count = 0
                for row_idx in range(2, sheet.max_row + 1):
                    phone_number = normalize_value(sheet.cell(row=row_idx, column=phone_col_idx).value)
                    if phone_number in statuses_to_update:
                        sheet.cell(row=row_idx, column=status_col_idx, value=statuses_to_update[phone_number])
                        updated_count += 1
                log_system(f"Batch update: Mapped {updated_count} statuses in memory.")
                wb.save(excel_file)
                wb.close()
                release_lock(excel_file)
                log_system(f"Batch update saved.")
                success = True
                return True
            except InvalidFileException:
                log_system(f"Error: '{excel_file}' invalid/corrupted. Batch update failed.")
                release_lock(excel_file)
                return False
            except Exception as e:
                log_system(f"Error during batch update attempt {attempt + 1}: {e}")
                logging.exception("Batch Update Traceback:")
                try: wb.close()
                except: pass
                if attempt < max_attempts - 1: 
                    time.sleep(retry_delay)
                else:
                    release_lock(excel_file)
    if not success: log_system(f"Batch update failed after {max_attempts} attempts.")
    return success

# --- Message Personalization (Modified for Custom Placeholders) ---
def parse_spintax(text: str) -> str:
    """Process spintax like [option1|option2] randomly."""
    pattern = re.compile(r"\[([^{}\[\]]+?)\]")
    while match := pattern.search(text):
        options = [opt.strip() for opt in match.group(1).split("|")]
        text = text[:match.start()] + random.choice(options) + text[match.end():]
    return text

def personalize_message(encoded_template: str, contact_details: Dict[str, Any], custom_placeholders: Dict[str, str]) -> str:
    """
    Decode, substitute standard & custom placeholders, process spintax, and re-encode.
    Args:
        encoded_template: The URL-encoded message template.
        contact_details: Dictionary representing the contact's row data from LIST sheet.
        custom_placeholders: Dictionary mapping {placeholder_keyword: list_column_header}.
    """
    if not encoded_template or pd.isna(encoded_template):
        return ""
    try:
        decoded_template = urllib.parse.unquote_plus(str(encoded_template))
    except Exception as e:
        logging.error(f"Error decoding template: {e}. Template: {encoded_template}")
        return ""

    # 1. Build the formatting dictionary including custom placeholders
    format_data = {}
    for keyword, list_header in custom_placeholders.items():
        # Get value from contact details using the header specified in PLACEHOLDER sheet
        format_data[keyword] = normalize_value(contact_details.get(list_header, ""))

    # 3. Perform formatting
    try:
        personalized = decoded_template.format(**format_data)
    except KeyError as e:
        placeholder_key = str(e).strip("'\"")
        if re.match(r"^{.+}$", decoded_template): # Check if template had placeholders
            log_system(f"Warning: Placeholder '{{{placeholder_key}}}' used in message template but not found in available data (standard or custom) for contact {contact_details.get(_LIST_COLS['phone'], 'N/A')}. Removing placeholder.")
        else: # Error is likely due to curly braces used for non-placeholder reasons
            log_system(f"Warning: Formatting error likely due to unexpected curly braces in message template for {contact_details.get(_LIST_COLS['phone'], 'N/A')}. Check template structure.")
        # Attempt to continue by removing the problematic key indication if it's simple
        # This is basic, might fail for complex templates. Better fix is usually correcting the template/data.
        personalized = decoded_template.replace(f"{{{placeholder_key}}}", "")
    except Exception as e:
        logging.error(f"Error formatting message: {e}. Template: {decoded_template}, Details: {contact_details}")
        personalized = decoded_template # Fallback

    # 4. Process Spintax and Re-encode
    spintax_processed = parse_spintax(personalized)
    return urllib.parse.quote_plus(spintax_processed)


# --- Excel Data Loader (Modified for Custom Placeholders) ---
class ExcelDataLoader:
    def __init__(self, excel_file: str) -> None:
        self.excel_file = excel_file
        self.sheets: Dict[str, pd.DataFrame] = {}
        self.settings: Dict[str, Any] = {}
        self.contacts_df: Optional[pd.DataFrame] = None
        self.messages_map: Dict[str, str] = {}
        self.docs_map: Dict[str, List[str]] = {}
        self.media_map: Dict[str, List[str]] = {}
        self.custom_placeholders: Dict[str, str] = {}
        self.browsers_map: Dict[str, List[str]] = {}

        # Create a temporary directory for Google Drive downloads for this session
        self.gdrive_download_cache = Path(tempfile.gettempdir()) / "wa_blaster_gdrive_downloads_cache"
        self.gdrive_download_cache.mkdir(parents=True, exist_ok=True)
        log_system(f"Google Drive download cache directory: {self.gdrive_download_cache}")

        self._load_data()

    def _load_settings_from_sheet(self) -> Dict[str, Any]:
        """Loads settings from the SETTINGS sheet."""
        settings_sheet_name = CONFIG["SHEETS"]["SETTINGS"]
        # Start with defaults
        current_settings = CONFIG["DEFAULT_SETTINGS"].copy()
        try:
            settings_df = pd.read_excel(self.excel_file, sheet_name=settings_sheet_name)
            loaded_settings = dict(zip(settings_df["Setting Name"], settings_df["Value"]))
            for key, value in loaded_settings.items():
                if key in current_settings: # Only update known settings from default
                    current_settings[key] = normalize_value(value)
                elif key and not pd.isna(key): # Warn about unknown settings
                    log_system(f"Warning: Unknown setting '{key}' in {settings_sheet_name} sheet ignored.")

            # Normalize specific settings
            current_settings[_SETTINGS_COLS["use_browser"]] = str(current_settings.get(_SETTINGS_COLS["use_browser"], "CHROME")).upper().strip()
            wd_ver_val = current_settings.get(_SETTINGS_COLS["wd_ver"])
            current_settings[_SETTINGS_COLS["wd_ver"]] = normalize_value(wd_ver_val) if wd_ver_val and normalize_value(wd_ver_val) else None
            custom_path_val = current_settings.get(_SETTINGS_COLS["custom_bsr_path"])
            current_settings[_SETTINGS_COLS["custom_bsr_path"]] = normalize_value(custom_path_val) if custom_path_val and normalize_value(custom_path_val) else None
            
            log_system(f"{settings_sheet_name} sheet processed.")
        except Exception as e:
            log_system(f"Warning: Could not load '{settings_sheet_name}' sheet: {e}. Using default settings.")
            # Ensure critical defaults are set if sheet loading fails
            current_settings = CONFIG["DEFAULT_SETTINGS"].copy() # Re-assign to ensure clean defaults
            current_settings[_SETTINGS_COLS["use_browser"]] = str(current_settings.get(_SETTINGS_COLS["use_browser"], "CHROME")).upper().strip()
            current_settings[_SETTINGS_COLS["wd_ver"]] = None
            current_settings[_SETTINGS_COLS["custom_bsr_path"]] = None
        return current_settings

    def _load_browsers_from_sheet(self) -> Dict[str, List[str]]:
        """Loads browser definitions from the BROWSER sheet."""
        browsers_map: Dict[str, List[str]] = {}
        browser_sheet_name = CONFIG["SHEETS"]["BROWSER"]
        browser_name_col = _BROWSER_COLS["name"]
        browser_path_cols = _BROWSER_COLS["paths"]
        try:
            browser_df = pd.read_excel(self.excel_file, sheet_name=browser_sheet_name, dtype=str)
            browser_df.dropna(subset=[browser_name_col], inplace=True)
            browser_df[browser_name_col] = browser_df[browser_name_col].str.upper().str.strip()

            for _, row in browser_df.iterrows():
                name = row[browser_name_col]
                if not name: continue

                paths = []
                for col in browser_path_cols:
                    if col in row and pd.notna(row[col]):
                        raw_path = str(row[col]).strip()
                        if raw_path:
                            expanded_path = os.path.expandvars(raw_path)
                            paths.append(expanded_path)
                if paths:
                    browsers_map[name] = paths
                else:
                    log_system(f"Warning: No paths defined for browser '{name}' in {browser_sheet_name} sheet.")
            
            if browsers_map:
                log_system(f"Loaded {len(browsers_map)} browser definitions from '{browser_sheet_name}': {', '.join(browsers_map.keys())}")
            else:
                log_system(f"Warning: No valid browser definitions found in '{browser_sheet_name}' sheet.")
        except FileNotFoundError:
            log_system(f"Warning: '{browser_sheet_name}' sheet not found. Browser path searching will be limited.")
        except Exception as e:
            log_system(f"Warning: Could not load '{browser_sheet_name}' sheet: {e}. Browser path searching disabled.")
        return browsers_map

    def _load_placeholders_from_sheet(self) -> Dict[str, str]:
        """Loads custom placeholders from the PLACEHOLDER sheet."""
        custom_placeholders: Dict[str, str] = {}
        placeholder_sheet_name = CONFIG["SHEETS"]["PLACEHOLDER"]
        placeholder_keyword_col = _PLACEHOLDER_COLS["keyword"]
        placeholder_header_col = _PLACEHOLDER_COLS["list_header"]
        try:
            placeholder_df = pd.read_excel(self.excel_file, sheet_name=placeholder_sheet_name)
            if placeholder_keyword_col not in placeholder_df.columns or \
               placeholder_header_col not in placeholder_df.columns:
                log_system(f"Warning: '{placeholder_sheet_name}' sheet missing required columns ('{placeholder_keyword_col}', '{placeholder_header_col}'). Custom placeholders disabled.")
            else:
                placeholder_df[placeholder_keyword_col] = placeholder_df[placeholder_keyword_col].astype(str).str.strip()
                placeholder_df[placeholder_header_col] = placeholder_df[placeholder_header_col].astype(str).str.strip()
                placeholder_df.dropna(subset=[placeholder_keyword_col, placeholder_header_col], inplace=True)
                placeholder_df = placeholder_df[placeholder_df[placeholder_keyword_col] != ""]

                custom_placeholders = dict(zip(placeholder_df[placeholder_keyword_col], placeholder_df[placeholder_header_col]))
                if custom_placeholders:
                    log_system(f"Loaded {len(custom_placeholders)} custom placeholders from '{placeholder_sheet_name}': {', '.join(custom_placeholders.keys())}")
                else:
                    log_system(f"No valid custom placeholders defined in '{placeholder_sheet_name}' sheet.")
        except FileNotFoundError:
            log_system(f"Optional sheet '{placeholder_sheet_name}' not found. Custom placeholders disabled.")
        except Exception as e:
            log_system(f"Warning: Could not load '{placeholder_sheet_name}' sheet: {e}. Custom placeholders disabled.")
        return custom_placeholders

    def _load_contacts_from_sheet(self, custom_placeholders: Dict[str, str]) -> pd.DataFrame:
        """Loads contacts from the LIST sheet, using custom placeholders for column setup."""
        contacts_df = pd.DataFrame()
        list_sheet_name = CONFIG["SHEETS"]["LIST"]
        list_converters = {
            _LIST_COLS["phone"]: str, _LIST_COLS["msg_code"]: str, _LIST_COLS["doc_code"]: str,
            _LIST_COLS["media_code"]: str, _LIST_COLS["status"]: str,
        }
        # Add converters for columns used by custom placeholders
        list_converters.update({header: str for header in custom_placeholders.values()})

        try:
            loaded_df = pd.read_excel(self.excel_file, sheet_name=list_sheet_name, converters=list_converters)
            for col in [_LIST_COLS["phone"], _LIST_COLS["msg_code"], _LIST_COLS["doc_code"], _LIST_COLS["media_code"]]:
                if col in loaded_df.columns:
                    loaded_df[col] = loaded_df[col].apply(normalize_value)
                else:
                    raise ValueError(f"Missing required column '{col}' in {list_sheet_name} sheet.")

            missing_custom_cols = []
            for list_header in custom_placeholders.values():
                if list_header in loaded_df.columns:
                    loaded_df[list_header] = loaded_df[list_header].apply(normalize_value)
                else:
                    missing_custom_cols.append(list_header)
            if missing_custom_cols:
                log_system(f"Warning: The following columns specified in '{CONFIG['SHEETS']['PLACEHOLDER']}' sheet were not found in '{list_sheet_name}' sheet: {', '.join(missing_custom_cols)}")

            status_col = _LIST_COLS["status"]
            if status_col not in loaded_df.columns:
                loaded_df[status_col] = _STATUS_VALS["PENDING"]
            else:
                loaded_df[status_col] = pd.to_numeric(loaded_df[status_col], errors='coerce')
                loaded_df[status_col] = loaded_df[status_col].fillna(_STATUS_VALS["PENDING"])
                loaded_df[status_col] = loaded_df[status_col].astype(int)
            
            contacts_df = loaded_df
            log_system(f"'{list_sheet_name}' sheet loaded: {len(contacts_df)} contacts.")
        except Exception as e:
            log_system(f"Error loading '{list_sheet_name}' sheet: {e}. No contacts will be processed.")
            # contacts_df remains an empty DataFrame initialized at the start of the method
        return contacts_df

    def _load_messages_from_sheet(self) -> Dict[str, str]:
        """Loads messages from the MSGS sheet."""
        messages_map: Dict[str, str] = {}
        msgs_sheet_name = CONFIG["SHEETS"]["MSGS"]
        msgs_code_col = CONFIG["COLUMNS"]["MSGS"]["msg_code"]
        msgs_msg_col = CONFIG["COLUMNS"]["MSGS"]["message"]
        try:
            msgs_df = pd.read_excel(self.excel_file, sheet_name=msgs_sheet_name, converters={msgs_code_col: str})
            msgs_df[msgs_code_col] = msgs_df[msgs_code_col].apply(normalize_value)
            # Ensure message column is treated as string, even if it contains numbers that pandas might auto-convert
            msgs_df[msgs_msg_col] = msgs_df[msgs_msg_col].astype(str) 
            messages_map = dict(zip(msgs_df[msgs_code_col], msgs_df[msgs_msg_col]))
            log_system(f"'{msgs_sheet_name}' sheet loaded: {len(messages_map)} messages.")
        except Exception as e:
            log_system(f"Warning: Could not load '{msgs_sheet_name}' sheet: {e}. Message functionality may be affected.")
        return messages_map

    def _load_file_mapping(self, sheet_key: str, col_config: Dict[str, Any]) -> Dict[str, List[str]]:
        mapping: Dict[str, List[str]] = {}
        sheet_name = CONFIG["SHEETS"][sheet_key]
        code_col = col_config["code"]
        file_cols = col_config["files"]
        
        try:
            df = pd.read_excel(self.excel_file, sheet_name=sheet_name, converters={code_col: str})
            df[code_col] = df[code_col].apply(normalize_value)
            df.dropna(subset=[code_col], inplace=True)

            for _, row in df.iterrows():
                key = row[code_col]
                if not key or key == '0':
                    continue

                raw_file_entries = [row.get(f_col) for f_col in file_cols if f_col in row and pd.notna(row.get(f_col))]
                
                valid_paths = []
                invalid_files_found = []

                for p_entry in raw_file_entries:
                    p = normalize_value(p_entry)
                    if p:
                        is_gdrive_link = "drive.google.com/" in p.lower() and \
                                        ("file/d/" in p.lower() or "/uc?" in p.lower() or "/open?" in p.lower() or "view?id=" in p.lower())

                        if is_gdrive_link:
                            log_system(f"Detected Google Drive link for code '{key}': {p}. Attempting download...")
                            
                            downloaded_path = self._download_gdrive_with_proper_naming(p, key)
                            
                            if downloaded_path and Path(downloaded_path).exists():
                                log_system(f"Successfully downloaded '{Path(downloaded_path).name}' for code '{key}': {downloaded_path}")
                                valid_paths.append(downloaded_path)
                            else:
                                log_system(f"Failed to download Google Drive file for code '{key}': {p}")
                                invalid_files_found.append(f"{p} (Download failed)")
                        
                        elif Path(p).exists() and Path(p).is_file():
                            valid_paths.append(str(Path(p).resolve()))
                        elif Path(p).exists() and not Path(p).is_file():
                            log_system(f"Warning: Path exists but is not a file for code '{key}': {p}")
                            invalid_files_found.append(f"{p} (Not a file)")
                        else:
                            invalid_files_found.append(f"{p} (Local file not found)")
                
                if invalid_files_found:
                    log_system(f"Warning: For code '{key}' in {sheet_name}, some files were problematic: {', '.join(invalid_files_found)}")
                
                mapping[key] = valid_paths
                if not valid_paths and invalid_files_found:
                    log_system(f"Note: Code '{key}' in {sheet_name} has no valid files after processing.")

            log_system(f"'{sheet_name}' sheet loaded: {len(mapping)} codes processed.")
            
        except FileNotFoundError:
            log_system(f"Warning: Sheet '{sheet_name}' not found in '{self.excel_file}'. {sheet_key} functionality disabled.")
        except Exception as e:
            log_system(f"Warning: Could not load '{sheet_name}' sheet from '{self.excel_file}': {e}. {sheet_key} functionality may be affected.")
            logging.exception(f"Traceback for _load_file_mapping {sheet_name}:")
        
        return mapping

    def _download_gdrive_with_proper_naming(self, url: str, code: str) -> str:
        """Download Google Drive file and ensure proper naming with extension"""
        cache_dir = Path(self.gdrive_download_cache)
        cache_dir.mkdir(exist_ok=True)
        
        # Clean up any existing .part files first
        self._cleanup_partial_downloads()
        
        file_id = self._extract_gdrive_file_id(url)
        if not file_id:
            log_system(f"Could not extract file ID from URL: {url}")
            return None
        
        # First, try to get file metadata to determine the proper filename and extension
        proper_filename, content_type = self._get_gdrive_file_metadata(file_id, code)
        
        # Try gdown download first
        log_system(f"Attempting gdown download for code '{code}'...")
        try:
            downloaded_path = gdown.download(
                url=f"https://drive.google.com/uc?id={file_id}",
                output=str(cache_dir) + "/",
                quiet=True,
                fuzzy=True
            )
            
            # Check if gdown succeeded
            if downloaded_path and Path(downloaded_path).exists() and not downloaded_path.endswith('.part'):
                # If gdown worked perfectly, just rename if needed
                final_path = self._ensure_proper_filename(downloaded_path, proper_filename, content_type, code)
                return final_path
                
        except Exception as e:
            log_system(f"gdown failed: {e}")
        
        # Handle .part files or failed downloads
        part_files = list(cache_dir.glob("*.part"))
        if part_files:
            log_system(f"Found {len(part_files)} .part files, processing...")
            
            # Get the largest .part file (most likely the one we want)
            largest_part = max(part_files, key=lambda f: f.stat().st_size)
            
            if largest_part.stat().st_size > 100:  # At least 100 bytes
                # Rename .part file to proper name with extension
                final_filename = proper_filename or f"{code}_{file_id}"
                final_path = cache_dir / final_filename
                
                try:
                    shutil.move(str(largest_part), str(final_path))
                    log_system(f"Successfully processed .part file to: {final_path}")
                    
                    # Clean up any remaining .part files
                    for part_file in part_files:
                        if part_file.exists():
                            part_file.unlink()
                    
                    return str(final_path)
                    
                except Exception as e:
                    log_system(f"Failed to process .part file: {e}")
        
        # If all else fails, try direct requests download
        log_system(f"Trying direct download for code '{code}'...")
        return self._direct_download_gdrive(file_id, code, proper_filename, content_type)

    def _get_gdrive_file_metadata(self, file_id: str, code: str) -> tuple:
        """Get file metadata from Google Drive to determine proper filename and type"""
        try:
            # Try to get file info from Google Drive API without authentication
            metadata_url = f"https://drive.google.com/file/d/{file_id}/view"
            
            response = requests.get(metadata_url, timeout=10)
            if response.status_code == 200:
                # Try to extract filename from page content
                content = response.text
                
                # Look for filename in various places
                filename_patterns = [
                    r'"title":"([^"]+)"',
                    r'<title>([^<]+) - Google Drive</title>',
                    r'"filename":"([^"]+)"',
                    r'data-filename="([^"]+)"'
                ]
                
                for pattern in filename_patterns:
                    match = re.search(pattern, content)
                    if match:
                        filename = match.group(1).strip()
                        if filename and filename != "Untitled":
                            log_system(f"Detected filename from metadata: {filename}")
                            return filename, None
                
                # Try to detect content type from page
                if 'application/pdf' in content or 'pdf' in content.lower():
                    return f"{code}.pdf", "application/pdf"
                elif 'image/' in content:
                    return f"{code}.jpg", "image/jpeg"  # Default to jpg for images
                elif 'text/' in content or 'document' in content.lower():
                    return f"{code}.txt", "text/plain"
                    
        except Exception as e:
            log_system(f"Could not get metadata for {file_id}: {e}")
        
        # Default fallback
        return f"{code}_{file_id}", None

    def _direct_download_gdrive(self, file_id: str, code: str, filename: str = None, content_type: str = None) -> str:
        """Direct download using requests as fallback"""
        cache_dir = Path(self.gdrive_download_cache)
        
        try:
            download_url = f"https://drive.google.com/uc?id={file_id}&export=download"
            
            with requests.Session() as session:
                response = session.get(download_url, stream=True, timeout=30)
                
                # Handle virus scan warning for large files
                if response.status_code == 200 and 'virus scan warning' in response.text.lower():
                    # Extract confirm token
                    confirm_token = None
                    for key, value in response.cookies.items():
                        if key.startswith('download_warning'):
                            confirm_token = value
                            break
                    
                    if confirm_token:
                        params = {'id': file_id, 'confirm': confirm_token}
                        response = session.get(download_url, params=params, stream=True, timeout=30)
                
                if response.status_code == 200:
                    # Determine filename
                    final_filename = filename or f"{code}_{file_id}"
                    
                    # Try to get filename from headers
                    if 'content-disposition' in response.headers:
                        cd = response.headers['content-disposition']
                        if 'filename=' in cd:
                            header_filename = cd.split('filename=')[1].strip('"\'')
                            if header_filename:
                                final_filename = header_filename
                    
                    # Ensure extension based on content type
                    if not Path(final_filename).suffix and content_type:
                        ext = mimetypes.guess_extension(content_type)
                        if ext:
                            final_filename += ext
                    
                    file_path = cache_dir / final_filename
                    
                    # Download file
                    with open(file_path, 'wb') as f:
                        for chunk in response.iter_content(chunk_size=8192):
                            if chunk:
                                f.write(chunk)
                    
                    if file_path.exists() and file_path.stat().st_size > 0:
                        log_system(f"Direct download successful: {file_path}")
                        return str(file_path)
                    
        except Exception as e:
            log_system(f"Direct download failed for {file_id}: {e}")
        
        return None

    def _ensure_proper_filename(self, file_path: str, desired_filename: str, content_type: str, code: str) -> str:
        """Ensure the downloaded file has the proper name and extension"""
        current_path = Path(file_path)
        
        if not current_path.exists():
            return file_path
        
        # If we have a desired filename, use it
        if desired_filename:
            new_path = current_path.parent / desired_filename
        else:
            # Try to detect file type and add appropriate extension
            new_name = current_path.name
            if not current_path.suffix:
                # No extension, try to detect
                extension = self._detect_file_extension(current_path, content_type)
                new_name = f"{code}_{current_path.name}{extension}"
                new_path = current_path.parent / new_name
            else:
                new_path = current_path
        
        # Rename if needed
        if new_path != current_path:
            try:
                shutil.move(str(current_path), str(new_path))
                log_system(f"Renamed file to: {new_path}")
                return str(new_path)
            except Exception as e:
                log_system(f"Could not rename file: {e}")
                return str(current_path)
        
        return str(current_path)

    def _detect_file_extension(self, file_path: Path, content_type: str = None) -> str:
        """Detect file extension based on content"""
        if content_type:
            ext = mimetypes.guess_extension(content_type)
            if ext:
                return ext
        
        # Read first few bytes to detect file type
        try:
            with open(file_path, 'rb') as f:
                header = f.read(32)
            
            # Common file signatures
            if header.startswith(b'%PDF'):
                return '.pdf'
            elif header.startswith(b'\xff\xd8\xff'):
                return '.jpg'
            elif header.startswith(b'\x89PNG'):
                return '.png'
            elif header.startswith(b'GIF8'):
                return '.gif'
            elif header.startswith(b'PK'):
                return '.zip'  # Could also be docx, xlsx, etc.
            elif b'<html' in header.lower() or b'<!doctype' in header.lower():
                return '.html'
            
        except Exception as e:
            log_system(f"Could not detect file type: {e}")
        
        return ''

    def _extract_gdrive_file_id(self, url: str) -> str:
        """Extract Google Drive file ID from URL"""
        patterns = [
            r'/file/d/([a-zA-Z0-9-_]+)',
            r'id=([a-zA-Z0-9-_]+)',
            r'/uc\?id=([a-zA-Z0-9-_]+)',
            r'/open\?id=([a-zA-Z0-9-_]+)'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, url)
            if match:
                return match.group(1)
        
        return None

    def _cleanup_partial_downloads(self):
        """Clean up any existing .part files from previous failed downloads"""
        try:
            cache_dir = Path(self.gdrive_download_cache)
            part_files = list(cache_dir.glob("*.part"))
            for part_file in part_files:
                try:
                    part_file.unlink()
                    log_system(f"Cleaned up partial download: {part_file}")
                except Exception as e:
                    log_system(f"Warning: Could not remove partial file {part_file}: {e}")
            
            if part_files:
                log_system(f"Cleaned up {len(part_files)} partial download files")
                
        except Exception as e:
            log_system(f"Warning: Error during cleanup of partial downloads: {e}")

    def _load_data(self) -> None:
        """Loads all data from the Excel file by calling helper methods."""
        try:
            self.settings = self._load_settings_from_sheet()
            self.browsers_map = self._load_browsers_from_sheet()
            self.custom_placeholders = self._load_placeholders_from_sheet()
            # Pass custom_placeholders to _load_contacts_from_sheet as it needs them
            self.contacts_df = self._load_contacts_from_sheet(self.custom_placeholders)
            self.messages_map = self._load_messages_from_sheet()
            self.docs_map = self._load_file_mapping("DOCS", CONFIG["COLUMNS"]["DOCS"])
            self.media_map = self._load_file_mapping("MEDIA", CONFIG["COLUMNS"]["MEDIA"])

        except Exception as e:
            logging.exception(f"Critical error during the overall data loading process: {e}")
            # Fallback initialization to ensure attributes exist
            self.settings = CONFIG["DEFAULT_SETTINGS"].copy()
            self.contacts_df = pd.DataFrame()
            self.messages_map = {}
            self.docs_map = {}
            self.media_map = {}
            self.custom_placeholders = {}
            self.browsers_map = {}
            log_system("ExcelDataLoader: Falling back to default/empty values due to critical loading error.")

    def _resolve_name(self, row: pd.Series) -> str:
        name_col = "_ResolvedName"
        raw_name = normalize_value(row.get(name_col, ""))
        return raw_name

    # --- Public Accessors ---
    def get_settings(self) -> Dict[str, Any]: return self.settings
    def get_contacts(self) -> pd.DataFrame: return self.contacts_df if self.contacts_df is not None else pd.DataFrame()
    def get_messages_map(self) -> Dict[str, str]: return self.messages_map
    def get_docs_map(self) -> Dict[str, List[str]]: return self.docs_map
    def get_media_map(self) -> Dict[str, List[str]]: return self.media_map
    def get_custom_placeholders(self) -> Dict[str, str]: return self.custom_placeholders
    def get_browsers_map(self) -> Dict[str, List[str]]: return self.browsers_map

# --- Browser Manager ---
class BrowserManager:
    def __init__(self, instance_id: int) -> None:
        self.driver: Optional[WebDriver] = None
        self.instance_id = instance_id
        self.user_data_path = get_persistent_temp_path(str(instance_id))
        self.browser_type: Optional[str] = None

    def get_latest_version():
        url = "https://googlechromelabs.github.io/chrome-for-testing/"
        response = requests.get(url)
        response.raise_for_status()
        tree = html.fromstring(response.content)
        xpath = "/html/body/div/table/tbody/tr[1]/td[1]/code"
        elements = tree.xpath(xpath)
        return elements[0].text()

    def _find_browser_executable(self, browser_name: str, browsers_map: Dict[str, List[str]]) -> Optional[str]:
        """Finds the first valid executable path for the given browser name."""
        browser_name_upper = browser_name.upper()
        if browser_name_upper not in browsers_map:
            log_system(f"Error: Browser '{browser_name}' not defined in BROWSER sheet.")
            return None

        possible_paths = browsers_map[browser_name_upper]
        log_system(f"Searching for {browser_name_upper} executable in: {possible_paths}")
        for path_str in possible_paths:
            path = Path(path_str)
            if path.exists() and path.is_file():
                log_system(f"Found {browser_name_upper} executable at: {path}")
                return str(path)

        log_system(f"Error: Could not find a valid executable for {browser_name_upper} in the specified paths.")
        return None

    def setup_browser(self,
                  headless: bool = False,
                  settings: Dict[str, Any] = {},
                  browsers_map: Dict[str, List[str]] = {}
                 ) -> Optional[WebDriver]:
        if self.driver:
            try:
                _ = self.driver.window_handles
                log_browser(self.instance_id, f"Browser instance already running.")
                return self.driver
            except WebDriverException:
                log_browser(self.instance_id, f"Browser instance crashed/closed. Re-initializing.")
                self.driver = None

        # --- Get common settings ---
        driver_version = settings.get(_SETTINGS_COLS["wd_ver"]) # WD_VER applies always
        selected_browser_name = settings.get(_SETTINGS_COLS["use_browser"], "CHROME").upper()

        browser_path = None
        browser_name_to_use = None # Will hold the final name ('CHROME', 'EDGE', etc.)

        # --- >> LOGIC BASED ON USE_BROWSER VALUE << ---
        if selected_browser_name == "CUSTOM":
            log_system("USE_BROWSER is set to CUSTOM. Reading CUSTOM_BSR_PATH...")
            custom_path_setting = settings.get(_SETTINGS_COLS["custom_bsr_path"])
            if not custom_path_setting:
                log_system(f"Error setup instance {self.instance_id}: USE_BROWSER is CUSTOM, but CUSTOM_BSR_PATH is empty in settings.")
                return None

            custom_path = Path(custom_path_setting)
            log_system(f"Checking custom browser path: {custom_path}")
            if custom_path.exists() and custom_path.is_file():
                log_system(f"Valid custom browser path found: {custom_path}")
                browser_path = str(custom_path)
                # Determine type for the custom path (using heuristic)
                filename = custom_path.name.lower()
                if "chrome.exe" in filename or "chrome" == filename: browser_name_to_use = "CHROME"
                elif "msedge.exe" in filename or "msedge" == filename: browser_name_to_use = "EDGE"
                elif "brave.exe" in filename or "brave" == filename or "brave-browser" == filename: browser_name_to_use = "BRAVE"
                elif "vivaldi.exe" in filename or "vivaldi" == filename: browser_name_to_use = "VIVALDI"
                elif "opera.exe" in filename or "opera" == filename: browser_name_to_use = "OPERA"
                else:
                    log_system(f"Warning: Could not reliably determine type for custom path '{filename}'. Assuming CHROME-compatible.")
                    browser_name_to_use = "CHROME" # Default assumption
                log_system(f"Using CUSTOM path. Determined type: {browser_name_to_use}")
            else:
                log_system(f"Error setup instance {self.instance_id}: Custom browser path '{custom_path}' is invalid or not found.")
                return None # Fail if CUSTOM is selected but path is bad

        else: # USE_BROWSER is not "CUSTOM" (e.g., "CHROME", "BRAVE", etc.)
            log_system(f"USE_BROWSER is '{selected_browser_name}'. Looking up in BROWSER sheet data...")
            if not browsers_map:
                log_system(f"Error setup instance {self.instance_id}: Browser definitions (BROWSER sheet) not loaded or empty, cannot find path for '{selected_browser_name}'.")
                return None

            # Find executable using the helper function and the selected name
            browser_path = self._find_browser_executable(selected_browser_name, browsers_map)
            if browser_path:
                browser_name_to_use = selected_browser_name # Use the name from the setting
            else:
                # Error logged in _find_browser_executable
                log_system(f"Failed setup for instance {self.instance_id}: Could not locate executable for '{selected_browser_name}' via BROWSER sheet.")
                return None

        # --- >> End of Conditional Logic << ---

        # --- Check if we finally have a valid path ---
        if not browser_path or not browser_name_to_use:
            # This case should ideally be caught above, but as a safeguard:
            log_system(f"Failed setup instance {self.instance_id}: No valid browser executable or type determined.")
            return None

        # --- Now, proceed with WebDriver setup using browser_path and browser_name_to_use ---
        self.browser_type = browser_name_to_use # Store the determined type
        service_args = None
        os.environ['WDM_LOG'] = str(logging.WARNING)
        if platform.system() == "Windows": service_args = ['--log-level=OFF']

        try:
            service = None
            options = None
            wdm_driver_version = driver_version if driver_version else self.get_latest_version()

            log_system(f"Configuring WebDriver for: {browser_name_to_use} (Path: {browser_path}, Driver Version: {wdm_driver_version})")

            # Map browser_name_to_use to WebDriver Managers and Options
            # (This mapping logic remains the same as the previous step)
            if browser_name_to_use == "CHROME":
                options = ChromeOptions()
                service = ChromeService(ChromeDriverManager(driver_version=wdm_driver_version).install(), service_args=service_args)
                self.browser_type_for_selenium = "chrome"
            elif browser_name_to_use == "EDGE":
                options = EdgeOptions()
                # options.binary_location = browser_path # Often needed, set below universally
                service = EdgeService(EdgeChromiumDriverManager(driver_version=wdm_driver_version).install(), service_args=service_args)
                self.browser_type_for_selenium = "edge"
            elif browser_name_to_use == "BRAVE":
                options = ChromeOptions()
                # options.binary_location = browser_path # Often needed, set below universally
                service = ChromeService(ChromeDriverManager(chrome_type=ChromeType.BRAVE, driver_version=wdm_driver_version).install(), service_args=service_args)
                self.browser_type_for_selenium = "chrome"
            elif browser_name_to_use in ["VIVALDI", "OPERA"]:
                options = ChromeOptions()
                # options.binary_location = browser_path # Often needed, set below universally
                log_system(f"Attempting to use standard ChromeDriver for {browser_name_to_use}")
                service = ChromeService(ChromeDriverManager(driver_version=wdm_driver_version).install(), service_args=service_args)
                self.browser_type_for_selenium = "chrome"
            else:
                log_system(f"Error: Unsupported browser type '{browser_name_to_use}' determined.")
                return None

            # --- Configure Options (Set binary location universally) ---
            options.add_experimental_option('excludeSwitches', ['enable-logging'])
            options.binary_location = browser_path # Set the final path here

            # ... (rest of the options configuration remains the same) ...
            options.add_argument(f"--user-data-dir={self.user_data_path}")
            options.add_argument(f"--remote-debugging-port={9222 + self.instance_id}")
            options.add_argument("--disable-extensions")
            options.add_argument("--disable-default-apps")
            options.add_argument("--disable-popup-blocking")
            options.add_argument("--no-first-run")
            options.add_argument("--no-service-autorun")
            options.add_argument("--password-store=basic")
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option('useAutomationExtension', False)
            options.add_argument(f"user-agent={settings.get('USER_AGENT')}")
            if headless:
                options.add_argument("--headless=new")
                options.add_argument("--window-size=1920,1080")

            log_system(f"Initializing {browser_name_to_use} instance {self.instance_id} (Headless: {headless})...")

            # --- Initialize WebDriver ---
            if self.browser_type_for_selenium == "chrome":
                self.driver = webdriver.Chrome(service=service, options=options)
            elif self.browser_type_for_selenium == "edge":
                self.driver = webdriver.Edge(service=service, options=options)

            log_system(f"{browser_name_to_use} instance {self.instance_id} initialized.")
            return self.driver

        # ... (Exception handling remains the same) ...
        except WebDriverException as e:
            # ... (Log specific errors)
            error_msg = str(e)
            if "This version of ChromeDriver only supports Chrome version" in error_msg \
            or "Current browser version is" in error_msg \
            or "cannot find" in error_msg.lower():
                log_browser(self.instance_id, f"Error init: WebDriver/Browser version mismatch or browser not found. Check WD_VER and paths (CUSTOM_BSR_PATH or BROWSER sheet) for {browser_name_to_use}.")
                log_browser(self.instance_id, f"Details: {error_msg}")
            else:
                log_browser(self.instance_id, f"WebDriver error initializing {browser_name_to_use}: {e}")
            logging.exception(f"WebDriver Init Traceback ({self.instance_id}):")
            self.driver = None
            return None
        except Exception as e:
            log_browser(self.instance_id, f"Unexpected error initializing instance {self.instance_id} ({browser_name_to_use}): {e}")
            logging.exception(f"General Init Traceback ({self.instance_id}):")
            self.driver = None
            return None

    def quit(self) -> None: # [source: 87-88]
        if self.driver:
            log_system(f"Quitting browser instance {self.instance_id}...")
            try:
                self.driver.quit()
            except Exception as e: log_system(f"Exception while quitting browser instance {self.instance_id}: {e}")
            finally:
                self.driver = None
                log_system(f"Browser instance {self.instance_id} quit.")


# --- WhatsApp Interaction Functions ---
def wait_for_element(driver: WebDriver, xpath: str, timeout: int = 10) -> Optional[Any]:
    try: return WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.XPATH, xpath)))
    except TimeoutException: return None
    except NoSuchElementException: return None
    except Exception: return None

def wait_for_clickable(driver: WebDriver, xpath: str, timeout: int = 10) -> Optional[Any]:
    try: return WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((By.XPATH, xpath)))
    except TimeoutException: return None
    except Exception: return None

def send_text_message(
    driver: WebDriver,
    phone_number: str,
    message_template: str,
    contact_details: Dict[str, Any],
    instance_id: int,
    settings: Dict[str, Any],
    custom_placeholders: Dict[str, str] # Added custom_placeholders arg
) -> str:
    """
    Sends a text message to the specified phone number using WhatsApp Web.

    Args:
        driver: The Selenium WebDriver instance.
        phone_number: The recipient's phone number.
        message_template: The template for the message.
        contact_details: A dictionary containing contact-specific information for placeholders.
        instance_id: The ID of the current instance.
        settings: A dictionary containing configuration settings, including XPath locators.
        custom_placeholders: A dictionary of custom placeholders and their values.

    Returns:
        A string indicating the status of the message ("SENT", "FAILED", or "INVALID").
    """
    status = "FAILED"
    if not message_template or pd.isna(message_template):
        return "INVALID"

    final_message_encoded = personalize_message(message_template, contact_details, custom_placeholders) # Pass custom_placeholders
    if not final_message_encoded:
        return "INVALID"

    base_url = "https://web.whatsapp.com"
    send_url = f"{base_url}/send?phone={phone_number}&text={final_message_encoded}&app_absent=0"

    try:
        driver.get(send_url)
        xpath_textbox = settings.get("XPATH_TEXT", CONFIG["DEFAULT_SETTINGS"]["XPATH_TEXT"])
        invalid_msg_text = settings.get("INVALID_MSG", CONFIG["DEFAULT_SETTINGS"]["INVALID_MSG"])
        xpath_invalid_popup_button = f"//div[contains(text(), \"{invalid_msg_text}\")]/ancestor::div[@role='dialog']//button"

        try:
            WebDriverWait(driver, 15).until(
                EC.any_of(
                    EC.presence_of_element_located((By.XPATH, xpath_textbox)),
                    EC.presence_of_element_located((By.XPATH, xpath_invalid_popup_button)),
                )
            )
        except TimeoutException:
            log_browser(instance_id, f"Timeout waiting for chat/popup for {phone_number}.")
            if "Scan QR code" in driver.page_source:
                log_browser(instance_id, "QR scan needed.")
            return "FAILED"

        invalid_popup = driver.find_elements(By.XPATH, xpath_invalid_popup_button)
        if invalid_popup:
            log_browser(instance_id, f"Invalid number {phone_number}.")
            try:
                invalid_popup[0].click()
            except Exception:
                pass
            return "INVALID"

        send_button_xpath = settings.get("XPATH_SEND", CONFIG["DEFAULT_SETTINGS"]["XPATH_SEND"])
        send_button = wait_for_clickable(driver, send_button_xpath, timeout=10)
        if not send_button:
            log_browser(instance_id, f"Send button not found for {phone_number}.")
            return "FAILED"

        time.sleep(random.uniform(0.5, 1.5))
        try:
            send_button.click()
            status = "SENT"
            time.sleep(random.uniform(1.0, 2.0))
        except (ElementClickInterceptedException, StaleElementReferenceException) as click_err:
            log_browser(instance_id, f"Click failed for send button ({phone_number}): {click_err}. Retrying JS...")
            try:
                time.sleep(1)
                send_button = wait_for_clickable(driver, send_button_xpath, timeout=5)
                if send_button:
                    driver.execute_script("arguments[0].click();", send_button)
                    status = "SENT"
                    time.sleep(random.uniform(1.0, 2.0))
                else:
                    log_browser(instance_id, f"Send button not found on retry.")
                    status = "FAILED"
            except Exception as retry_err:
                log_browser(instance_id, f"Error sending (retry click): {retry_err}")
                status = "FAILED"
        except Exception as e:
            log_browser(instance_id, f"Unexpected error sending: {e}")
            status = "FAILED"
    except WebDriverException as e:
        log_browser(instance_id, f"WebDriver error sending text: {e}")
        status = "FAILED"
        if "disconnected" in str(e):
            log_browser(instance_id, "Browser disconnected.")
    except Exception as e:
        log_browser(instance_id, f"General error sending text: {e}")
        status = "FAILED"
    return status

def attach_and_send_files(driver: WebDriver,
                          phone_number: str,
                          file_paths: List[str],
                          file_type: str,
                          instance_id: int,
                          settings: Dict[str, Any]) -> bool:
    if not file_paths:
        return True 

    chat_url = f"https://web.whatsapp.com/send?phone={phone_number}"

    try:
        # Navigate to the chat if not already there
        if phone_number not in driver.current_url:
            log_browser(instance_id, f"Navigating to chat for {phone_number} to attach {file_type}...")
            driver.get(chat_url)
            xpath_text = settings.get("XPATH_TEXT", CONFIG["DEFAULT_SETTINGS"]["XPATH_TEXT"])
            if not wait_for_element(driver, xpath_text, timeout=10):
                log_browser(instance_id, f"Failed to open chat for {phone_number}.")
                return False

        time.sleep(random.uniform(1.0, 2.0))

        # Click the attach (paperclip) button
        attach_xpath = settings.get("XPATH_ATTACH", CONFIG["DEFAULT_SETTINGS"]["XPATH_ATTACH"])
        attach_button = wait_for_clickable(driver, attach_xpath, timeout=10)
        if not attach_button:
            log_browser(instance_id, f"Attach button not found for {phone_number}.")
            return False
        try:
            attach_button.click()
        except Exception:
            driver.execute_script("arguments[0].click();", attach_button)
        time.sleep(1)

        # Force all file inputs to be visible
        def force_visible_inputs():
            driver.execute_script("""
                const inputs = document.querySelectorAll('input[type="file"]');
                for (const input of inputs) {
                    input.style.display = 'block';
                    input.style.visibility = 'visible';
                    input.style.width = '1px';
                    input.style.height = '1px';
                    input.style.opacity = 1;
                }
            """)

        force_visible_inputs()
        time.sleep(0.5)

        # Select correct input and fallback icon
        if file_type == "DOCS":
            xpath_input = '//input[@type="file" and @accept="*"]'
            xpath_option = settings.get("XPATH_DOCS", CONFIG["DEFAULT_SETTINGS"]["XPATH_DOCS"])
        elif file_type == "MEDIA":
            xpath_input = '//input[@type="file" and contains(@accept, "image")]'
            xpath_option = settings.get("XPATH_MEDIA", CONFIG["DEFAULT_SETTINGS"]["XPATH_MEDIA"])
        else:
            log_browser(instance_id, f"Unknown file type '{file_type}'")
            return False

        # Try to locate input element
        try:
            file_input = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, xpath_input))
            )
        except TimeoutException:
            log_browser(instance_id, f"{file_type} input not found after forcing visibility for {phone_number}. Retrying...")

            option_button = wait_for_clickable(driver, xpath_option, timeout=5)
            if option_button:
                try:
                    option_button.click()
                    time.sleep(1)
                    force_visible_inputs()
                    file_input = WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located((By.XPATH, xpath_input))
                    )
                except Exception as retry_err:
                    log_browser(instance_id, f"Retry failed: Could not expose file input for {file_type} ({phone_number}): {retry_err}")
                    return False
            else:
                log_browser(instance_id, f"Retry failed: {file_type} icon button not found.")
                return False

        # Check if any valid files exist
        absolute_paths = [str(Path(p).resolve()) for p in file_paths if Path(p).exists()]
        if not absolute_paths:
            log_browser(instance_id, f"No valid files to send to {phone_number}. Skipping.")
            return True

        # Send files
        try:
            file_input.send_keys("\n".join(absolute_paths))
            log_browser(instance_id, f"Sent {len(absolute_paths)} {file_type} file(s) to {phone_number}.")
        except Exception as send_err:
            log_browser(instance_id, f"Error sending file paths to input for {phone_number}: {send_err}")
            return False

        # Click the final "Send" button
        send_xpath = settings.get("XPATH_ASEND", CONFIG["DEFAULT_SETTINGS"]["XPATH_ASEND"])
        send_button = wait_for_clickable(driver, send_xpath, timeout=30)
        if not send_button:
            log_browser(instance_id, f"Send button (after attachment) not found for {phone_number}.")
            return False
        try:
            send_button.click()
        except Exception:
            driver.execute_script("arguments[0].click();", send_button)
        time.sleep(random.uniform(1.0, 2.5))
        return True

    except Exception as e:
        log_browser(instance_id, f"General error sending {file_type} to {phone_number}: {e}")
        logging.exception("Attachment error:")
        return False

# --- Core Contact Processing Logic (Modified) ---
def process_contact(
    driver: WebDriver,
    contact: pd.Series,
    messages_map: Dict[str, str],
    docs_map: Dict[str, List[str]],
    media_map: Dict[str, List[str]],
    instance_id: int,
    settings: Dict[str, Any],
    stop_event: threading.Event,
    final_statuses: Dict[str, int],
    status_update_lock: threading.Lock,
    custom_placeholders: Dict[str, str],
) -> None:
    """
    Processes a single contact, using custom placeholders & updating final_statuses dict.
    """
    if stop_event.is_set():
        return

    phone = normalize_value(contact.get(_LIST_COLS["phone"]))
    resolved_name = contact.get(_LIST_COLS["resolved_name"], "")

    if not phone:
        log_browser(instance_id, f"Skipping missing phone (Index: {contact.name}).")
        return

    current_status = _STATUS_VALS["RETRY"]

    try:
        # Extract codes
        msg_code = normalize_value(contact.get(_LIST_COLS["msg_code"]))
        doc_code = normalize_value(contact.get(_LIST_COLS["doc_code"]))
        media_code = normalize_value(contact.get(_LIST_COLS["media_code"]))

        contact_details = contact.to_dict()

        # Look up templates/files
        message_template = messages_map.get(msg_code) if msg_code != "0" else None
        doc_files = docs_map.get(doc_code)    if doc_code  != "0" else None
        media_files = media_map.get(media_code) if media_code!= "0" else None

        has_message = bool(msg_code != "0" and message_template)
        has_docs = bool(doc_code != "0" and doc_files)
        has_media = bool(media_code != "0" and media_files)

        # Validation
        if msg_code != "0" and msg_code not in messages_map:
            log_browser(instance_id, f"Invalid Msg Code {msg_code} for {phone}.")
            current_status = _STATUS_VALS["INVALID"]
        elif doc_code != "0" and doc_code not in docs_map:
            log_browser(instance_id, f"Invalid Doc Code {doc_code} for {phone}.")
            current_status = _STATUS_VALS["INVALID"]
        elif media_code != "0" and media_code not in media_map:
            log_browser(instance_id, f"Invalid Media Code {media_code} for {phone}.")
            current_status = _STATUS_VALS["INVALID"]
        elif not (has_message or has_docs or has_media):
            log_browser(instance_id, f"No actions for {phone}. Marking SENT.")
            current_status = _STATUS_VALS["SENT"]

        # If already invalid/sent, record and exit
        if current_status != _STATUS_VALS["RETRY"]:
            with status_update_lock:
                final_statuses[phone] = current_status
            return

        # --- Perform Actions ---
        message_sent_status = "SKIPPED"
        docs_sent  = True
        media_sent = True

        # 1) Send text
        if has_message:
            log_browser(instance_id,
                        f"Processing {phone}... Sending msg (Code: {msg_code})")
            min_t = float(settings.get("MIN_TIMER", "2.0"))
            max_t = float(settings.get("MAX_TIMER", "5.0"))
            time.sleep(random.uniform(min_t, max_t))

            message_sent_status = send_text_message(
                driver, phone, message_template,
                contact_details, instance_id,
                settings, custom_placeholders
            )

            if stop_event.is_set():
                return

            if message_sent_status == "INVALID":
                current_status = _STATUS_VALS["INVALID"]
                docs_sent = False
                media_sent = False
            elif message_sent_status == "FAILED":
                current_status = _STATUS_VALS["RETRY"]
                docs_sent = False
                media_sent = False

        # 2) Optionally wait before attachments
        elif has_docs or has_media:
            time.sleep(random.uniform(1.0, 2.0))

        # 3) Attach documents
        if has_docs and docs_sent:
            valid_doc_files = docs_map.get(doc_code, [])
            if valid_doc_files:
                log_browser(instance_id,
                            f"Attaching {len(valid_doc_files)} docs for {phone} (Code: {doc_code})...")
                docs_sent = attach_and_send_files(
                    driver, phone, valid_doc_files, "DOCS",
                    instance_id, settings
                )
            else:
                log_browser(instance_id,
                            f"Skipping docs: Code '{doc_code}' has no valid files.")
                docs_sent = True

            if stop_event.is_set():
                return
            if not docs_sent:
                current_status = _STATUS_VALS["RETRY"]
                media_sent = False

        # 4) Attach media
        if has_media and media_sent:
            valid_media_files = media_map.get(media_code, [])
            if valid_media_files:
                log_browser(instance_id,
                            f"Attaching {len(valid_media_files)} media for {phone} (Code: {media_code})...")
                media_sent = attach_and_send_files(
                    driver, phone, valid_media_files, "MEDIA",
                    instance_id, settings
                )
            else:
                log_browser(instance_id,
                            f"Skipping media: Code '{media_code}' has no valid files.")
                media_sent = True

            if stop_event.is_set():
                return
            if not media_sent:
                current_status = _STATUS_VALS["RETRY"]

        # 5) Final status decision
        if (message_sent_status != "FAILED"
            and docs_sent
            and media_sent
            and current_status == _STATUS_VALS["RETRY"]):
            log_browser(instance_id, f"Successfully processed {phone}.")
            current_status = _STATUS_VALS["SENT"]
        elif current_status == _STATUS_VALS["RETRY"]:
            log_browser(instance_id, f"Marking {phone} for retry.")

        with status_update_lock:
            final_statuses[phone] = current_status

    except WebDriverException as wd_err:
        log_browser(instance_id,
                    f"WebDriverException processing {phone}: {wd_err}. Retrying.")
        logging.exception(f"WD Traceback ({instance_id}, {phone}):")
        current_status = _STATUS_VALS["RETRY"]
        with status_update_lock:
            final_statuses[phone] = current_status

    except Exception as e:
        log_browser(instance_id,
                    f"Unexpected error processing {phone}: {e}. Retrying.")
        logging.exception(f"Processing Traceback ({instance_id}, {phone}):")
        current_status = _STATUS_VALS["RETRY"]
        with status_update_lock:
            final_statuses[phone] = current_status

# --- PySide6 GUI Implementation ---
class WhatsAppBlasterGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.excel_file: Optional[str] = None
        self.settings: Dict[str, Any] = {}
        self.browsers_map: Dict[str, List[str]] = {} # <-- Store browser map here too
        self.headless_mode = False
        # ... (rest of __init__ remains the same) ...
        self.signals = Signals()
        global global_signals
        global_signals = self.signals
        self.final_statuses: Dict[str, int] = {}
        self.status_update_lock = threading.Lock()
        self.total_processed_count = 0
        self.browser_manager_1 = BrowserManager(instance_id=1)
        self.browser_manager_2 = BrowserManager(instance_id=2)
        self.processing_thread: Optional[threading.Thread] = None
        self.stop_event = threading.Event()
        self.init_ui()
        self.connect_signals()
        log_system("Application initialized.")

    def init_ui(self):
        self.setWindowTitle("WhatsApp Blaster - PySide6 Batch Update")
        self.setFixedSize(550, 730)

        main_widget = QWidget(); self.setCentralWidget(main_widget)
        main_layout = QVBoxLayout(main_widget)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(10)

        # Title 
        title = QLabel("WhatsApp Blaster")
        title_font = QFont("Arial", 16, QFont.Bold)
        title.setFont(title_font)
        title.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title)

        # File Section 
        file_group = QWidget()
        file_layout = QGridLayout(file_group)
        file_layout.addWidget(QLabel("Excel File:"), 0, 0)
        self.excel_path_display = QLineEdit()
        self.excel_path_display.setReadOnly(True)
        self.excel_path_display.setPlaceholderText("No file selected...")
        file_layout.addWidget(self.excel_path_display, 0, 1, 1, 2)
        self.btn_import = QPushButton("Import Excel")
        file_layout.addWidget(self.btn_import, 1, 1)
        self.btn_template = QPushButton("Download Template")
        file_layout.addWidget(self.btn_template, 1, 2)
        main_layout.addWidget(file_group)

        # Control Buttons [source: 154-155]
        control_group = QWidget(); control_layout = QGridLayout(control_group)
        self.btn_launch = QPushButton("Launch WA Web (Login)")
        control_layout.addWidget(self.btn_launch, 0, 0)
        self.btn_quit = QPushButton("Quit Browsers")
        control_layout.addWidget(self.btn_quit, 0, 1)
        self.btn_run = QPushButton("RUN")
        self.btn_run.setStyleSheet("background-color: #4CAF50; color: white;")
        control_layout.addWidget(self.btn_run, 1, 0)
        self.btn_stop = QPushButton("STOP")
        self.btn_stop.setEnabled(False)
        self.btn_stop.setStyleSheet("background-color: #f44336; color: white;")
        control_layout.addWidget(self.btn_stop, 1, 1)
        self.chk_headless = QCheckBox("Run Headless")
        control_layout.addWidget(self.chk_headless, 2, 0, 1, 2, alignment=Qt.AlignCenter)
        main_layout.addWidget(control_group)

        # Progress Bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setFormat("%v / %m Contacts")
        main_layout.addWidget(self.progress_bar)

        # Utility Buttons 
        util_group = QWidget()
        util_layout = QHBoxLayout(util_group)
        self.btn_encoder = QPushButton("Encoder")
        self.btn_decoder = QPushButton("Decoder")
        self.btn_clear_cache = QPushButton("Clear Browser Cache")
        util_layout.addWidget(self.btn_encoder)
        util_layout.addWidget(self.btn_decoder)
        util_layout.addWidget(self.btn_clear_cache)
        main_layout.addWidget(util_group)

        # Log Areas [source: 157-158]
        log_group = QWidget()
        log_layout = QVBoxLayout(log_group)
        log_layout.setSpacing(5)
        log_layout.addWidget(QLabel("System Logs:"))
        self.sys_log = QTextEdit()
        self.sys_log.setReadOnly(True)
        self.sys_log.setFixedHeight(100)
        log_layout.addWidget(self.sys_log)
        log_layout.addWidget(QLabel("Browser 1 Logs:"))
        self.b1_log = QTextEdit()
        self.b1_log.setReadOnly(True)
        self.b1_log.setFixedHeight(100)
        log_layout.addWidget(self.b1_log)
        log_layout.addWidget(QLabel("Browser 2 Logs:"))
        self.b2_log = QTextEdit(); self.b2_log.setReadOnly(True)
        self.b2_log.setFixedHeight(100)
        log_layout.addWidget(self.b2_log)
        main_layout.addWidget(log_group)
        main_layout.addStretch()

        # Footer 
        footer = QLabel("Developed by qt3000@hw.ac.uk")
        footer.setAlignment(Qt.AlignCenter)
        footer_font = QFont("Arial", 8)
        footer.setFont(footer_font)
        main_layout.addWidget(footer)
        self.set_style()

    def set_style(self):
        self.setStyleSheet("""
            QMainWindow { background-color: #333333; }
            QLabel { color: white; font: 10pt Arial; }
            QPushButton { background-color: #555555; color: white; border: 1px solid #666666; border-radius: 4px; padding: 5px 10px; min-width: 80px; }
            QPushButton:hover { background-color: #666666; }
            QPushButton:pressed { background-color: #444444; }
            QPushButton:disabled { background-color: #444444; color: #888888; border: 1px solid #555555;}
            QLineEdit { background-color: #444444; color: white; border: 1px solid #555555; border-radius: 4px; padding: 4px; }
            QTextEdit { background-color: #2B2B2B; color: #A9B7C6; border: 1px solid #444444; border-radius: 4px; font-family: Consolas, monospace; }
            QCheckBox { color: white; spacing: 5px; }
            QCheckBox::indicator { width: 16px; height: 16px; border-radius: 3px; }
            QCheckBox::indicator:unchecked { background-color: #555; border: 1px solid #666; }
            QCheckBox::indicator:checked { background-color: #4CAF50; border: 1px solid #66AF70; }
            QProgressBar { border: 1px solid grey; border-radius: 5px; text-align: center; color: white; background-color: #444444; }
            QProgressBar::chunk { background-color: #4CAF50; width: 10px; margin: 0.5px; border-radius: 2px; }
        """)

    def connect_signals(self):
        # Button Clicks
        self.btn_import.clicked.connect(self.import_excel)
        self.btn_template.clicked.connect(self.download_template)
        self.btn_launch.clicked.connect(self.launch_browsers)
        self.btn_quit.clicked.connect(self.quit_browsers)
        self.btn_run.clicked.connect(self.run_blaster)
        self.btn_stop.clicked.connect(self.stop_blaster)
        self.btn_encoder.clicked.connect(lambda: self.open_coder_window('encode'))
        self.btn_decoder.clicked.connect(lambda: self.open_coder_window('decode'))
        self.btn_clear_cache.clicked.connect(self.delete_temp_folders)
        self.chk_headless.toggled.connect(self.toggle_headless)
        # Custom Signals
        self.signals.log_system_signal.connect(self.append_sys_log)
        self.signals.log_browser1_signal.connect(self.append_b1_log)
        self.signals.log_browser2_signal.connect(self.append_b2_log)
        self.signals.processing_started.connect(self.on_processing_started)
        self.signals.processing_stopped.connect(self.on_processing_stopped)
        self.signals.update_progress.connect(self.update_progress_bar)

    def append_sys_log(self, text: str):
        self.sys_log.append(text)
        self.sys_log.verticalScrollBar().setValue(self.sys_log.verticalScrollBar().maximum())

    def append_b1_log(self, text: str):
        self.b1_log.append(text)
        self.b1_log.verticalScrollBar().setValue(self.b1_log.verticalScrollBar().maximum())

    def append_b2_log(self, text: str):
        self.b2_log.append(text)
        self.b2_log.verticalScrollBar().setValue(self.b2_log.verticalScrollBar().maximum())

    # Progress Bar Slot
    def update_progress_bar(self, current_value: int, total_value: int):
         if total_value > 0: # Avoid division by zero if no contacts
              self.progress_bar.setMaximum(total_value)
              self.progress_bar.setValue(current_value)
              self.progress_bar.setFormat(f"{current_value} / {total_value} Contacts")
         else:
              self.progress_bar.setMaximum(1) # Prevent invalid range
              self.progress_bar.setValue(0)
              self.progress_bar.setFormat("0 / 0 Contacts")
    
    # GUI Action Slots
    def import_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Excel File", "", "Excel Files (*.xlsx *.xlsm)")
        if path:
            self.excel_file = path
            self.excel_path_display.setText(path)
            log_system(f"Selected Excel file: {path}")
            if not self.load_settings():
                self.excel_file = None
                self.excel_path_display.clear()

    def load_settings(self) -> bool:
        if not self.excel_file:
            log_system("Cannot load settings: No Excel file selected.")
            return False
        try:
            log_system(f"Loading data and settings from {self.excel_file}...")
            loader = ExcelDataLoader(self.excel_file)
            self.settings = loader.get_settings()
            self.browsers_map = loader.get_browsers_map()
            # Perform checks after loading
            if loader.get_contacts().empty:
                log_system("Warning: LIST sheet empty or failed to load during settings check.")
            if not self.settings:
                log_system("SETTINGS sheet missing or failed, using defaults.")
            if not self.browsers_map:
                log_system("Warning: BROWSER sheet failed to load or is empty. Browser setup might fail.")
            log_system("Settings and browser data loaded successfully.")
            return True
        except Exception as e:
            QMessageBox.critical(self, "Error Loading Excel", f"Failed to load data/settings from Excel:\n{e}")
            log_system(f"Critical error loading Excel data: {e}")
            self.settings = {}
            self.browsers_map = {}
            return False

    def download_template(self):
        template_filename = "Template.xlsx"
        try:
            if getattr(sys, 'frozen', False):
                base_path = Path(sys._MEIPASS)
            else:
                base_path = Path(__file__).resolve().parent

            template_path = base_path / template_filename

            if not template_path.exists():
                QMessageBox.critical(self, "Error", f"Template '{template_filename}' not found.")
                log_system(f"Template not found: {template_path}")
                return
            destination, _ = QFileDialog.getSaveFileName(self, "Save Template As", template_filename, "Excel Files (*.xlsx)")
            if destination:
                shutil.copyfile(str(template_path), destination)
                QMessageBox.information(self, "Success", f"Template saved to:\n{destination}"); log_system(f"Template saved: {destination}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save template: {e}")
            log_system(f"Error saving template: {e}")
    
    def launch_browsers(self):
        if not self.excel_file:
            QMessageBox.warning(self, "Settings Needed", "Please import an Excel file first.")
            return
        # Ensure settings AND browser map are loaded
        if not self.settings or not self.browsers_map:
            log_system("Attempting to load settings and browser data before launching browsers...")
            if not self.load_settings():
                 QMessageBox.critical(self, "Error", "Failed to load settings/browser data. Cannot launch.")
                 return
        if not self.browsers_map: # Specific check if map is still empty
            QMessageBox.critical(self, "Browser Error", "Browser definitions not loaded. Check BROWSER sheet. Cannot launch.")
            return

        log_system("Launching browser instances for WhatsApp Web login...")
        log_system("Please scan the QR code in each browser window if prompted.")
        QMessageBox.information(self, "Launch Browsers", "Attempting to launch browsers.\nPlease check for new browser windows and scan the WhatsApp QR code if needed.")

        # Pass settings AND browsers_map to the launch thread
        def launch(instance_id, settings, manager, browsers_map_arg): # Add browsers_map_arg
            # Pass browsers_map to setup_browser
            driver = manager.setup_browser(headless=False, settings=settings, browsers_map=browsers_map_arg)
            if driver:
                try:
                    current_url = ""
                    try: current_url = driver.current_url
                    except WebDriverException: log_browser(instance_id, "Browser seems to have closed."); return
                    if "web.whatsapp.com" not in current_url: driver.get("https://web.whatsapp.com")
                    log_browser(instance_id, "Browser launched. Ready for QR scan or already logged in.")
                except WebDriverException as e: log_browser(instance_id, f"Error navigating to WhatsApp Web: {e}")
            else: log_browser(instance_id, "Failed to launch browser.")

        # Start threads, passing the maps
        threading.Thread(target=launch, args=(1, self.settings.copy(), self.browser_manager_1, self.browsers_map.copy()), daemon=True, name="LaunchThread-1").start()
        threading.Thread(target=launch, args=(2, self.settings.copy(), self.browser_manager_2, self.browsers_map.copy()), daemon=True, name="LaunchThread-2").start()
    
    def quit_browsers(self):
        log_system("Quitting browser instances...")
        threading.Thread(target=self.browser_manager_1.quit, daemon=True).start()
        threading.Thread(target=self.browser_manager_2.quit, daemon=True).start()
        log_system("Browser quit commands issued.")

    def run_blaster(self):
        if not self.excel_file:
            QMessageBox.warning(self, "File Needed", "Import Excel first.")
            return
        if not self.settings or not self.browsers_map: # Ensure base settings are loaded
            log_system("Attempting to load settings and browser data before run...")
            if not self.load_settings():
                 QMessageBox.critical(self, "Error", "Failed to load settings/browser data. Cannot run.")
                 return
        if not self.browsers_map:
            QMessageBox.critical(self, "Browser Error", "Browser definitions not loaded. Check BROWSER sheet. Cannot run.")
            return

        if self.processing_thread and self.processing_thread.is_alive():
            QMessageBox.warning(self, "Running", "Processing active.")
            return

        log_system("--- Starting Blaster ---")
        self.stop_event.clear()
        self.final_statuses.clear()
        self.total_processed_count = 0

        # Pass the core settings and browser map. _processing_runner will handle loading run-specific data.
        self.processing_thread = threading.Thread(
            target=self._processing_runner,
            args=(self.excel_file, self.settings.copy(), self.browsers_map.copy(), self.headless_mode),
            daemon=True, name="BlasterRunner"
        )
        self.processing_thread.start()
        self.signals.processing_started.emit()

    def _execute_processing_phase(self,
                                 phase_name: str,
                                 contacts_for_phase: pd.DataFrame,
                                 messages_map: Dict[str, str],
                                 docs_map: Dict[str, List[str]],
                                 media_map: Dict[str, List[str]],
                                 driver1: Optional[WebDriver],
                                 driver2: Optional[WebDriver],
                                 settings: Dict[str, Any],
                                 custom_placeholders: Dict[str, str]) -> None:
        """
        Executes a single processing phase (e.g., Initial, Retry).
        """
        if contacts_for_phase.empty:
            log_system(f"{phase_name} run skipped: No contacts for this phase.")
            return

        total_to_process_for_phase = len(contacts_for_phase)
        log_system(f"{phase_name} run: {total_to_process_for_phase} contacts.")
        self.signals.update_progress.emit(0, total_to_process_for_phase)

        # Ensure drivers are valid before proceeding with the phase
        # Note: The decision to re-validate/re-setup drivers before *each* phase
        # (initial vs retry) is up to the calling logic in _processing_runner.
        # This function assumes drivers passed to it are ready or None.

        if not driver1 and not driver2:
            log_system(f"Cannot start {phase_name} phase: No active browser drivers.")
            # Mark all contacts in this phase for retry if no drivers are available
            with self.status_update_lock:
                for _, contact_row in contacts_for_phase.iterrows():
                    phone = normalize_value(contact_row.get(_LIST_COLS["phone"]))
                    if phone:
                        self.final_statuses[phone] = _STATUS_VALS["RETRY"]
            self._save_pending_updates() # Save the retry statuses
            return

        self.final_statuses.clear() # Clear for the current phase
        self.total_processed_count = 0 # Reset for the current phase

        contacts_list = [row for _, row in contacts_for_phase.iterrows()]

        self._run_phase(contacts_list, messages_map, docs_map, media_map,
                        driver1, driver2, settings, phase_name,
                        total_to_process_for_phase, custom_placeholders)

        log_system(f"Saving {phase_name} results...")
        self._save_pending_updates()
        if self.stop_event.is_set():
            log_system(f"Stopped during {phase_name} run.")

    def _processing_runner(self, excel_file: str, settings: Dict[str,Any], browsers_map: Dict[str,List[str]], headless: bool):
        """Main runner function executed in a separate thread."""
        loader = None
        driver1: Optional[WebDriver] = None
        driver2: Optional[WebDriver] = None

        try:
            log_system("Loading data for run...")
            loader = ExcelDataLoader(excel_file) # Load all data fresh for the run
            
            # Load all data needed for the entire run (initial + retry)
            contacts_df_full = loader.get_contacts()
            messages_map = loader.get_messages_map()
            docs_map = loader.get_docs_map()
            media_map = loader.get_media_map()
            custom_placeholders = loader.get_custom_placeholders()
            # Settings and browsers_map are passed in, but if you want them fresh from Excel too for each run:
            # settings = loader.get_settings() # Caution: This would override GUI chosen headless state etc.
            # browsers_map = loader.get_browsers_map()


            if contacts_df_full.empty:
                log_system("Run cancelled: LIST sheet in Excel is empty or failed to load.")
                QMessageBox.warning(None, "No Contacts", "The LIST sheet in your Excel file is empty or could not be loaded.")
                self.signals.processing_stopped.emit()
                return

            status_pending, status_retry = _STATUS_VALS["PENDING"], _STATUS_VALS["RETRY"]
            status_col = _LIST_COLS["status"]

            # --- Setup Browsers Once ---
            log_system("Setting up browsers for the run...")
            # Ensure browsers are quit before starting a new run, if they were left open from a previous action.
            # This is a good practice to ensure a clean state.
            self.browser_manager_1.quit()
            self.browser_manager_2.quit()
            time.sleep(1) # Give a moment for browsers to close

            driver1 = self.browser_manager_1.setup_browser(headless, settings, browsers_map)
            driver2 = self.browser_manager_2.setup_browser(headless, settings, browsers_map)

            if not driver1 and not driver2:
                log_system("Failed to start ANY browser drivers. Aborting run.")
                QMessageBox.critical(None, "Browser Error", "Could not start any browser instances. Please check settings and browser installations.")
                self.signals.processing_stopped.emit()
                return
            if not driver1: log_system("Warning: Browser Instance 1 failed to start.")
            if not driver2: log_system("Warning: Browser Instance 2 failed to start.")

            # --- Initial Processing Phase ---
            initial_contacts_df = contacts_df_full[contacts_df_full[status_col].isin([status_pending, status_retry])].copy()
            self._execute_processing_phase("Initial", initial_contacts_df,
                                           messages_map, docs_map, media_map,
                                           driver1, driver2, settings, custom_placeholders)

            if self.stop_event.is_set():
                log_system("Processing stopped after initial phase.")
                self.signals.processing_stopped.emit()
                return

            # --- Retry Processing Phase ---
            log_system("Reloading contact statuses for retry phase...")
            # It's crucial to reload the Excel to get the statuses that were just updated by perform_batch_update
            # This ensures we're retrying based on the latest saved state.
            loader_for_retry = ExcelDataLoader(excel_file)
            contacts_df_for_retry_full = loader_for_retry.get_contacts() # Get fresh contacts including updated statuses

            if contacts_df_for_retry_full.empty:
                log_system("Retry phase skipped: LIST sheet in Excel is empty or failed to load for retry.")
            else:
                retry_contacts_df = contacts_df_for_retry_full[contacts_df_for_retry_full[status_col] == status_retry].copy()
                
                # Re-validate drivers before retry phase. They might have crashed.
                # No need to fully re-setup unless _get_valid_driver indicates they are gone.
                # If a driver crashed, it will be None.
                current_driver1 = self._get_valid_driver(self.browser_manager_1)
                current_driver2 = self._get_valid_driver(self.browser_manager_2)

                # Optional: If you want to attempt re-setup if a driver crashed:
                if driver1 and not current_driver1:
                    log_system("Browser 1 seems to have disconnected. Attempting to restart for retry phase...")
                    current_driver1 = self.browser_manager_1.setup_browser(headless, settings, browsers_map)
                if driver2 and not current_driver2:
                    log_system("Browser 2 seems to have disconnected. Attempting to restart for retry phase...")
                    current_driver2 = self.browser_manager_2.setup_browser(headless, settings, browsers_map)

                self._execute_processing_phase("Retry", retry_contacts_df,
                                               messages_map, docs_map, media_map,
                                               current_driver1, current_driver2, settings, custom_placeholders)

            log_system("--- Blaster Run Finished ---")

        except Exception as e:
            log_system(f"Critical error in processing runner: {e}")
            logging.exception("Runner Traceback:")
            # Attempt to save any statuses that might have been collected before the error
            self._save_pending_updates()
        finally:
            # Ensure browsers are quit eventually, e.g. if user stops or run finishes
            # self.browser_manager_1.quit() # Or quit them at the very end of the application
            # self.browser_manager_2.quit()
            self.signals.processing_stopped.emit()

    def _run_phase(self, 
                   contacts_list: List[pd.Series],
                   messages_map: Dict,
                   docs_map: Dict,
                   media_map: Dict,
                   driver1: Optional[WebDriver],
                   driver2: Optional[WebDriver],
                   settings: Dict,
                   phase_name: str,
                   total_for_phase: int,
                   custom_placeholders: Dict[str, str]):

        """Runs processing phase, passing custom_placeholders to workers."""
        if not contacts_list: log_system(f"{phase_name} skipped: No contacts."); return
        q1, q2 = queue.Queue(), queue.Queue()
        for i, contact in enumerate(contacts_list):
            assigned=False
            phone = contact.get(_LIST_COLS['phone'], 'Unk')

            if i % 2 == 0 and driver1:
                q1.put(contact)
                assigned = True
            elif driver2:
                q2.put(contact)
                assigned = True
            elif driver1:
                q1.put(contact)
                assigned = True
            if not assigned:
                log_system(f"Err assign: {phone}")
            with self.status_update_lock:
                self.final_statuses[phone] = _STATUS_VALS["RETRY"]
        log_system(f"{phase_name}: Q1={q1.qsize()}, Q2={q2.qsize()}.")

        def browser_worker(instance_id, driver, contact_q): # Passes custom_placeholders to process_contact
            if not driver: return
            log_browser(instance_id, f"{phase_name} worker started.")
            while not contact_q.empty() and not self.stop_event.is_set():
                contact = None # Define contact outside try for except blocks
                try:
                    contact = contact_q.get_nowait()
                    process_contact(driver, contact, messages_map, docs_map, media_map, instance_id, settings, self.stop_event, self.final_statuses, self.status_update_lock, custom_placeholders) # Pass custom_placeholders
                    contact_q.task_done()
                    with self.status_update_lock: self.total_processed_count += 1; current_count = self.total_processed_count
                    self.signals.update_progress.emit(current_count, total_for_phase)
                except queue.Empty: break
                except WebDriverException as e:
                    phone = contact.get(_LIST_COLS['phone'], 'Unk') if contact is not None else 'Unk'
                    log_browser(instance_id, f"WD Exc {phase_name} for {phone}: {e}. Retrying.")
                    with self.status_update_lock: self.final_statuses[phone] = _STATUS_VALS["RETRY"]
                    with self.status_update_lock: self.total_processed_count += 1; current_count = self.total_processed_count
                    self.signals.update_progress.emit(current_count, total_for_phase)
                    if "disconnected" in str(e) or "fail" in str(e): log_browser(instance_id, "Browser disconnected. Worker stop."); break
                except Exception as e:
                    phone = contact.get(_LIST_COLS['phone'], 'Unk') if contact is not None else 'Unk'
                    log_browser(instance_id, f"Unexpected error {phase_name} for {phone}: {e}"); logging.exception(f"Worker Traceback ({instance_id}, {phase_name}):")
                    with self.status_update_lock: self.final_statuses[phone] = _STATUS_VALS["RETRY"]
                    with self.status_update_lock: self.total_processed_count += 1; current_count = self.total_processed_count
                    self.signals.update_progress.emit(current_count, total_for_phase)
            log_browser(instance_id, f"{phase_name} worker finished.")

        thread1 = threading.Thread(target=browser_worker, args=(1, driver1, q1), daemon=True, name=f"Worker-1-{phase_name}") if driver1 else None
        thread2 = threading.Thread(target=browser_worker, args=(2, driver2, q2), daemon=True, name=f"Worker-2-{phase_name}") if driver2 else None
        if thread1: thread1.start(); 
        if thread2: thread2.start()
        if thread1: thread1.join()
        if thread2: thread2.join()
        log_system(f"{phase_name} phase complete.")

    def _get_valid_driver(self, manager: BrowserManager) -> Optional[WebDriver]:
        if manager.driver:
            try:
                _ = manager.driver.window_handles
                return manager.driver
            except WebDriverException:
                log_browser(manager.instance_id, "Driver disconnected.")
                manager.driver = None
                return None
        return None

    def _save_pending_updates(self):
        if not self.excel_file:
            log_system("Cannot save updates: Excel file path not set.")
            return

        with self.status_update_lock:
            statuses_to_save = self.final_statuses.copy()

        if not statuses_to_save:
            log_system("No pending status updates to save.")
            return
        
        log_system(f"Attempting to save {len(statuses_to_save)} status updates...")
        
        save_thread = threading.Thread(target=perform_batch_update, args=(self.excel_file, statuses_to_save), daemon=True, name="BatchUpdateThread")
        save_thread.start()
        save_thread.join(timeout=30.0)

        if save_thread.is_alive():
            log_system("Warning: Batch update is taking a long time.")
            QMessageBox.warning(self, "Save Operation", "Saving Excel file is taking longer than expected.")

    def stop_blaster(self):
        if self.processing_thread and self.processing_thread.is_alive():
            log_system("--- Sending STOP signal ---")
            self.stop_event.set()
            self.btn_stop.setEnabled(False); self.btn_stop.setText("Stopping...") # 
            # The runner thread will call _save_pending_updates on stop
        else:
            log_system("Stop requested but no processing thread is active.")

    def on_processing_started(self):
        self.progress_bar.setValue(0)
        self.progress_bar.setFormat("0 / 0 Contacts") # Reset progress
        self.btn_run.setEnabled(False)
        self.btn_stop.setEnabled(True)
        self.btn_stop.setText("STOP")
        self.btn_launch.setEnabled(False)
        self.btn_import.setEnabled(False)
        self.btn_template.setEnabled(False)
        self.btn_clear_cache.setEnabled(False)
        log_system("GUI Controls updated: Processing started.")

    def on_processing_stopped(self):
        self.processing_thread = None
        self.btn_run.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self.btn_stop.setText("STOP")
        self.btn_launch.setEnabled(True)
        self.btn_import.setEnabled(True)
        self.btn_template.setEnabled(True)
        self.btn_clear_cache.setEnabled(True)
        log_system("GUI Controls updated: Processing stopped/finished.")

    def toggle_headless(self, state):
        self.headless_mode = bool(state)
        log_system(f"Headless mode {'enabled' if self.headless_mode else 'disabled'}.")

    def open_coder_window(self, mode):
        coder_win = QDialog(self)
        coder_win.setWindowTitle(f"Message {mode.capitalize()}r")
        coder_win.setFixedSize(400, 350); coder_win.setStyleSheet(self.styleSheet())

        layout = QVBoxLayout(coder_win)
        input_label = QLabel(f"Input ({'Raw Text' if mode == 'encode' else 'URL Encoded Text'}):")
        layout.addWidget(input_label); input_text = QTextEdit()
        layout.addWidget(input_text)
        output_label = QLabel(f"Output ({'URL Encoded' if mode == 'encode' else 'Decoded Text'}):")
        layout.addWidget(output_label); output_text = QTextEdit()
        output_text.setReadOnly(True); layout.addWidget(output_text)

        button_box = QHBoxLayout()
        btn = QPushButton(f"{mode.capitalize()} Message")
        btn.clicked.connect(lambda: self.handle_codec(mode, input_text, output_text))

        copy_btn = QPushButton("Copy Output")
        copy_btn.clicked.connect(lambda: self.copy_to_clipboard(output_text))
        close_btn = QPushButton("Close"); close_btn.clicked.connect(coder_win.accept)

        button_box.addWidget(btn); button_box.addWidget(copy_btn)
        button_box.addStretch(); button_box.addWidget(close_btn)
        layout.addLayout(button_box)
        coder_win.exec()

    def handle_codec(self, mode, input_widget, output_widget):
        msg = input_widget.toPlainText().strip()
        if not msg: output_widget.clear(); return
        try:
            if mode == 'encode':
                result = urllib.parse.quote_plus(msg)
            else:
                result = urllib.parse.unquote_plus(msg)
            output_widget.setPlainText(result)
        except Exception as e:
            QMessageBox.critical(self, "Codec Error", f"Processing failed: {str(e)}"); log_system(f"Codec error ({mode}): {e}")

    def copy_to_clipboard(self, text_widget):
        clipboard = QApplication.clipboard()
        text_to_copy = text_widget.toPlainText()

        if text_to_copy:
            clipboard.setText(text_to_copy)
            log_system("Output copied to clipboard.")

    def delete_temp_folders(self):
        temp_dir = Path(tempfile.gettempdir())
        folders_to_delete = [f"whatsapp_blaster_data_{i}" for i in [1, 2]]
        folders_to_delete.append("wa_blaster_gdrive_downloads_cache")

        deleted_count, error_count, not_found_count = 0, 0, 0
        reply = QMessageBox.question(self, "Confirm Delete", "Delete cached browser data?\nThis might require QR scan again.", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No)
        
        if reply == QMessageBox.StandardButton.Yes:
            log_system("Attempting to delete browser data folders...")
            for folder in folders_to_delete:
                folder_path = temp_dir / folder
                if folder_path.exists() and folder_path.is_dir():
                    try:
                        shutil.rmtree(folder_path)
                        log_system(f"Deleted: {folder_path}")
                        deleted_count += 1
                    except Exception as e:
                        log_system(f"Error deleting {folder_path}: {e}")
                        QMessageBox.critical(self, "Deletion Error", f"Could not delete:\n{folder_path}\nError: {e}")
                        error_count += 1
                elif folder_path.exists():
                    log_system(f"Path exists but not a directory: {folder_path}")
                    error_count += 1
                else:
                    log_system(f"Folder not found: {folder_path}")
                    not_found_count +=1

            summary = f"Deletion complete.\nDeleted: {deleted_count}\nFailed: {error_count}\nNot Found: {not_found_count}"
            QMessageBox.information(self, "Deletion Complete", summary); log_system(f"Folder deletion summary: {summary}")

        else:
            log_system("Browser data deletion cancelled.")

    def closeEvent(self, event: QCloseEvent):
        log_system("--- Closing requested ---"); allow_close = True
        if self.processing_thread and self.processing_thread.is_alive():
            reply = QMessageBox.question(self, "Confirm Exit", "Processing active. Exit now?\nWill try to save progress.", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.No:
                log_system("Close cancelled.")
                event.ignore()
                allow_close = False
            else:
                log_system("Stopping active processing for exit...")
                self.stop_event.set()
        if allow_close:
            log_system("Cleaning up...")
            self._save_pending_updates()
            self.quit_browsers(); time.sleep(0.5)
            log_system("Cleanup finished. Exiting.")
            event.accept()

if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO, 
        format="%(asctime)s - %(levelname)s - [%(threadName)s] - %(message)s",
        handlers=[logging.FileHandler("whatsapp_blaster.log", encoding='utf-8'),
        logging.StreamHandler()]
    )
    app = QApplication(sys.argv)
    window = WhatsAppBlasterGUI()
    window.show()
    sys.exit(app.exec())
